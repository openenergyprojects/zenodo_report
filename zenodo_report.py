#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Oct 12 12:08:00 2023
"""
import openpyxl
import requests
import os
import re
import json
import logging
from colorlog import ColoredFormatter

# Configure logging
LOG_LEVEL = logging.DEBUG  # Change to desired log level (DEBUG, INFO, WARNING, ERROR, CRITICAL)
LOG_FORMAT = "%(log_color)s%(levelname)-8s%(reset)s %(white)s%(message)s"
formatter = ColoredFormatter(LOG_FORMAT)

handler = logging.StreamHandler()
handler.setFormatter(formatter)

logger = logging.getLogger("zenodo_report")
logger.addHandler(handler)
logger.setLevel(LOG_LEVEL)

# Global parameters
OVERWRITE_OPEN_ACCESS_LINK = False  # Set to True to overwrite existing "Open Access link"
POPULATE_DOI_COLUMN = True  # Set to False to skip populating the DOI column
CURL_UA = "Mozilla/5.0 (Linux; Android 10; SM-G996U Build/QP1A.190711.020; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Mobile Safari/537.36"

def search_zenodo_by_title(title):
    """Search Zenodo for a record by title."""
    zenodo_api_url = "https://zenodo.org/api/records"
    params = {"q": f'title:"{title}"'}
    response = requests.get(zenodo_api_url, params=params)
    if response.status_code == 200:
        data = response.json()
        if data.get("hits", {}).get("total", 0) == 1:
            recid = data["hits"]["hits"][0]["id"]
            return recid, None
        elif data.get("hits", {}).get("total", 0) > 1:
            return None, "Multiple matches found"
    return None, "No match found"

def search_zenodo_by_doi(doi):
    """Search Zenodo for a record by DOI."""
    zenodo_api_url = "https://zenodo.org/api/records"
    params = {"q": f'doi:{doi}'}
    response = requests.get(zenodo_api_url, params=params)
    if response.status_code == 200:
        data = response.json()
        if data.get("hits", {}).get("total", 0) > 0:
            recid = data["hits"]["hits"][0]["id"]
            return recid
    return None

def extract_doi_from_bibliographic_data(bibliographic_data):
    """Extract DOI from the bibliographic data field."""
    match = re.search(r"DOI\s*:\s*([\S]+)", bibliographic_data)
    if match:
        return match.group(1).replace(" ", "")
    return None

def update_xlsx_with_zenodo_links(xlsx_file, sheet_names):
    """Update the XLSX file with Zenodo open access links, DOIs, and author emails for multiple sheets."""
    try:
        workbook = openpyxl.load_workbook(xlsx_file)
    except Exception as e:
        logger.error(f"Error loading workbook: {e}")
        return

    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in the XLSX file.")
            continue

        sheet = workbook[sheet_name]

        # Find headers and their column indices (headers are at row 3)
        headers = {cell.value: idx for idx, cell in enumerate(sheet[3], start=1)}
        if "NO." not in headers or "TITLE " not in headers or "BIBLIOGRAPHIC DATA" not in headers:
            logger.error(f"Required headers ('NO.', 'TITLE ', 'BIBLIOGRAPHIC DATA') not found in sheet '{sheet_name}'.")
            continue

        if "Open Access link" not in headers:
            headers["Open Access link"] = len(headers) + 1
            sheet.cell(row=3, column=headers["Open Access link"]).value = "Open Access link"

        if "DOI" not in headers:
            headers["DOI"] = 23  # Column W corresponds to index 23
            sheet.cell(row=3, column=headers["DOI"]).value = "DOI"

        if "link_as_text" not in headers:
            headers["link_as_text"] = 24  # Column X corresponds to index 24
            sheet.cell(row=3, column=headers["link_as_text"]).value = "link_as_text"

        if "author_email" not in headers:
            headers["author_email"] = 25  # Column Y corresponds to index 25
            sheet.cell(row=3, column=headers["author_email"]).value = "author_email"

        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
            no_cell = row[headers["NO."] - 1]
            title_cell = row[headers["TITLE "] - 1]
            bibliographic_data_cell = row[headers["BIBLIOGRAPHIC DATA"] - 1]
            open_access_link_cell = row[headers["Open Access link"] - 1]
            doi_cell = row[headers["DOI"] - 1]
            link_as_text_cell = row[headers["link_as_text"] - 1]
            author_email_cell = row[headers["author_email"] - 1]

            entry_id = str(no_cell.value).strip()

            # Skip rows that do not start with a number or are not the title row
            if not entry_id.isdigit() and entry_id != "NO.":
                continue

            title = str(title_cell.value).strip() if title_cell.value else ""
            bibliographic_data = str(bibliographic_data_cell.value).strip() if bibliographic_data_cell.value else ""

            # Skip if Zenodo link, DOI column, and link_as_text are all populated
            if open_access_link_cell.value and doi_cell.value and link_as_text_cell.value:
                log_with_context(f"Skipping entry as all required fields are populated: {title}", sheet_name, entry_id, level=logging.INFO)
                continue

            if not title:
                log_with_context(f"Skipping row with missing title: {row}", sheet_name, entry_id, level=logging.WARNING)
                continue

            log_with_context(f"Processing entry: {title}", sheet_name, entry_id, indent_level=0)

            # Extract link_as_text from the "LINK" column if enabled
            if link_as_text_cell.value is None and "LINK" in headers:
                links_cell = row[headers["LINK"] - 1]
                if links_cell.hyperlink:
                    link_as_text_cell.value = links_cell.hyperlink.target
                else:
                    link_as_text_cell.value = links_cell.value if links_cell.value else ""

            # Check if "Open Access link" should be updated
            if not open_access_link_cell.value or OVERWRITE_OPEN_ACCESS_LINK:
                log_with_context(f"Searching Zenodo for title: {title}", sheet_name, entry_id, indent_level=1)
                recid, message = search_zenodo_by_title(title)

                if recid:
                    zenodo_link = f"https://zenodo.org/records/{recid}"
                    log_with_context(f"Found Zenodo record: {zenodo_link}", sheet_name, entry_id, indent_level=2)
                    open_access_link_cell.value = zenodo_link

                    # Attempt to extract DOI from bibliographic data if available
                    if bibliographic_data:
                        doi = extract_doi_from_bibliographic_data(bibliographic_data)

                    # Use DOI if found, otherwise fallback to recid
                    if doi:
                        log_with_context(f"DOI extracted: {doi}", sheet_name, entry_id, indent_level=2)
                        doi_cell.value = f"https://doi.org/{doi}"
                    else:
                        log_with_context(f"No DOI found, using recid as fallback.", sheet_name, entry_id, indent_level=2)
                        doi_cell.value = f"https://doi.org/{recid}"
                else:
                    log_with_context(f"Zenodo search failed: {message}", sheet_name, entry_id, level=logging.WARNING, indent_level=1)
                    if bibliographic_data:
                        doi = extract_doi_from_bibliographic_data(bibliographic_data)
                        if doi:
                            log_with_context(f"Extracted DOI: {doi}", sheet_name, entry_id, indent_level=2)
                            recid = search_zenodo_by_doi(doi)
                            if recid:
                                zenodo_link = f"https://zenodo.org/records/{recid}"
                                log_with_context(f"Found Zenodo record for DOI: {zenodo_link}", sheet_name, entry_id, indent_level=2)
                                open_access_link_cell.value = zenodo_link
                                doi_cell.value = f"https://doi.org/{doi}"
                            else:
                                log_with_context(f"No Zenodo record found for DOI: {doi}", sheet_name, entry_id, level=logging.WARNING, indent_level=2)
                                doi_cell.value = f"https://doi.org/{doi}"
                        else:
                            log_with_context(f"No DOI found in bibliographic data.", sheet_name, entry_id, level=logging.WARNING, indent_level=2)
                    else:
                        log_with_context(f"No bibliographic data available.", sheet_name, entry_id, level=logging.WARNING, indent_level=1)
            else:
                log_with_context(f"Open Access link already populated, skipping update.", sheet_name, entry_id, indent_level=1)
                # Check if DOI is already populated
                if doi_cell.value is None and bibliographic_data:
                    doi = extract_doi_from_bibliographic_data(bibliographic_data)
                    if doi:
                        log_with_context(f"Extracted DOI: {doi}", sheet_name, entry_id, indent_level=2)
                        doi_cell.value = f"https://doi.org/{doi}"
                    else:
                        log_with_context(f"No DOI found in bibliographic data.", sheet_name, entry_id, level=logging.WARNING, indent_level=2)
                else:
                    log_with_context(f"DOI already populated, skipping update.", sheet_name, entry_id, indent_level=2)

            # Populate author_email column if not already populated
            if not author_email_cell.value and link_as_text_cell.value:
                author_email = extract_ucy_author_email(link_as_text_cell.value)
                if author_email:
                    author_email_cell.value = author_email

            # Save the workbook after processing each row
            try:
                workbook.save(xlsx_file)
                log_with_context(f"Workbook saved successfully after processing row.", sheet_name, entry_id, indent_level=0)
            except Exception as e:
                log_with_context(f"Error saving workbook: {e}", sheet_name, entry_id, level=logging.ERROR, indent_level=0)

    # Save the updated workbook
    try:
        workbook.save(xlsx_file)
        logger.info(f"Workbook saved successfully: {xlsx_file}")
    except Exception as e:
        logger.error(f"Error saving workbook: {e}")

    # At the end of the process, create a text file to prepare emails for authors with papers without entries in Zenodo.
    def prepare_email_file(xlsx_file, sheet_names, output_file):
        """Prepare a text file with emails for authors whose papers lack Zenodo entries."""
        try:
            workbook = openpyxl.load_workbook(xlsx_file)
        except Exception as e:
            logger.error(f"Error loading workbook: {e}")
            return

        author_papers = {}

        for sheet_name in sheet_names:
            if sheet_name not in workbook.sheetnames:
                logger.error(f"Sheet '{sheet_name}' not found in the XLSX file.")
                continue

            sheet = workbook[sheet_name]
            headers = {cell.value: idx for idx, cell in enumerate(sheet[3], start=1)}

            if "TITLE " not in headers or "Open Access link" not in headers or "author_email" not in headers:
                logger.error(f"Required headers ('TITLE ', 'Open Access link', 'author_email') not found in sheet '{sheet_name}'.")
                continue

            for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
                title_cell = row[headers["TITLE "] - 1]
                open_access_link_cell = row[headers["Open Access link"] - 1]
                link_as_text_cell = row[headers["link_as_text"] - 1]
                author_email_cell = row[headers["author_email"] - 1]
                doi_cell = row[headers["DOI"] - 1] if "DOI" in headers else None
                year_cell = row[headers["PUBLICATION YEAR"] - 1] if "PUBLICATION YEAR" in headers else None
                month_cell = row[headers["PUBLICATION MONTH"] - 1] if "PUBLICATION MONTH" in headers else None
                co_authors_cell = row[headers["AUTHORS"] - 1] if "AUTHORS" in headers else None

                title = str(title_cell.value).strip() if title_cell.value else ""
                open_access_link = str(open_access_link_cell.value).strip() if open_access_link_cell.value else ""
                link_as_text = str(link_as_text_cell.value).strip() if link_as_text_cell.value else ""
                author_email = str(author_email_cell.value).strip() if author_email_cell.value else ""
                doi = str(doi_cell.value).strip() if doi_cell and doi_cell.value else ""
                year = str(year_cell.value).strip() if year_cell and year_cell.value else ""
                month = str(month_cell.value).strip() if month_cell and month_cell.value else ""
                co_authors = str(co_authors_cell.value).strip() if co_authors_cell and co_authors_cell.value else ""

                if not open_access_link and author_email:
                    if author_email not in author_papers:
                        author_papers[author_email] = []
                    author_papers[author_email].append({
                        "title": title,
                        "doi": doi,
                        "year": year,
                        "month": month,
                        "co_authors": co_authors,
                        "link": link_as_text #open_access_link
                    })

        with open(output_file, "w") as email_file:
            email_file.write("Emails for authors with papers lacking Zenodo entries:\n\n")

            for author_email, papers in author_papers.items():
                email_file.write(f"Author: {author_email}\n")
                for paper in papers:
                    email_file.write(f"  Title: {paper['title']}\n")
                    email_file.write(f"  DOI: {paper['doi']}\n")
                    email_file.write(f"  Year: {paper['year']}\n")
                    email_file.write(f"  Month: {paper['month']}\n")
                    email_file.write(f"  Co-Authors: {paper['co_authors']}\n")
                    email_file.write(f"  Link: {paper['link']}\n\n")

        logger.info(f"Email preparation file created: {output_file}")

    # Call the function at the end of the update_xlsx_with_zenodo_links function
    output_email_file = "authors_without_zenodo_entries.txt"
    prepare_email_file(xlsx_file, sheet_names, output_email_file)

def extract_ucy_author_email(link):
    """Extract the email address of the author with ucy.ac.cy domain from the given link."""
    logger.debug(f"Starting email extraction for link: {link}")

    if "ieee.org" in link:
        logger.debug("Link contains 'ieee.org', proceeding to fetch metadata.")
        json = fetch_ieee_metadata(link)
        if not json:
            logger.debug("No JSON metadata fetched from IEEE link.")
            return None

        logger.debug(f"JSON metadata fetched: {json}")
        email = extract_email_from_ieee_json(json)
        if email:
            logger.debug(f"Email extracted from JSON metadata: {email}")
            return email

        logger.debug("No email found in JSON metadata, checking for specific author exceptions.")
        # Handle exceptions for specific authors
        if "Thomas Parisini" in json:
            logger.debug("Found 'Thomas Parisini' in JSON, returning exception email.")
            return "parisini.thomas@ucy.ac.cy"
        if "Alessandro Astolfi" in json:
            logger.debug("Found 'Alessandro Astolfi' in JSON, returning exception email.")
            return "astolfi.alessandro@ucy.ac.cy"

    logger.debug("Link does not contain 'ieee.org' or no email found, returning None.")
    return None

def fetch_ieee_metadata(link):
    """Fetch metadata from an IEEE link."""
    logger.debug(f"Fetching IEEE metadata for link: {link}")
    json = None
    try:
        response = requests.get(link, headers={"User-Agent": CURL_UA})
        logger.debug(f"HTTP response status code: {response.status_code}")
        if response.status_code == 200:
            json = extract_ieee_json(response.text)
            logger.debug(f"Extracted JSON from response: {json}")
        else:
            logger.debug(f"Failed to fetch metadata, status code: {response.status_code}")
    except Exception as e:
        logger.error(f"Error fetching IEEE metadata: {e}")
    return json

def extract_ieee_json(html):
    """Extract JSON metadata from IEEE HTML."""
    logger.debug("Extracting JSON metadata from HTML.")
    # Updated regex to match the JSON metadata structure in the provided HTML sample
    match = re.search(r'xplGlobal\.document\.metadata\s*=\s*(\{.*?\})\s*;', html)
    if match:
        logger.debug("JSON metadata found in HTML.")
        return json.loads(match.group(1))
    logger.debug("No JSON metadata found in HTML.")
    return None

def extract_email_from_ieee_json(json):
    """Extract the email address of the first author with ucy.ac.cy domain."""
    logger.debug("Extracting email from JSON metadata.")
    for author in json.get("authors", []):
        logger.debug(f"Processing author: {author}")
        for affiliation in author.get("affiliation", []):
            logger.debug(f"Checking affiliation: {affiliation}")
            if "ucy.ac.cy" in affiliation:
                email = f"{author['lastName'].lower()}.{author['firstName'].lower()}@ucy.ac.cy"
                logger.debug(f"Found email: {email}")
                return email

    logger.debug("No email found with 'ucy.ac.cy' domain. Attempting to construct email from names.")
    # Attempt to construct email from names of authors with "University of Cyprus" in their affiliation
    for author in json.get("authors", []):
        for affiliation in author.get("affiliation", []):
            if "University of Cyprus" in affiliation:
                first_name = author['firstName'].split()[0].lower()
                last_name = author['lastName'].split()[0].lower()
                email = f"{last_name}.{first_name}@ucy.ac.cy"
                logger.debug(f"Constructed email: {email}")
                return email

    logger.debug("No authors with 'University of Cyprus' affiliation found. Checking for specific exceptions.")
    # Handle exceptions for specific authors
    for author in json.get("authors", []):
        if author['firstName'] == "Thomas" and author['lastName'] == "Parisini":
            logger.debug("Found 'Thomas Parisini', returning exception email.")
            return "parisini.thomas@ucy.ac.cy"
        if author['lastName'] == "Astolfi":
            logger.debug("Found 'Astolfi', returning exception email.")
            return "astolfi.alessandro@ucy.ac.cy"

    logger.debug("No solution found for email extraction.")
    return None

def log_with_context(message, sheet_name, entry_id, level=logging.INFO, indent_level=0):
    """Log a message with context (sheet name and entry ID) and indentation."""
    indent = "   " * indent_level
    context = f"[Sheet: {sheet_name}, Entry ID: {entry_id}]"
    if level == logging.DEBUG:
        logger.debug(f"{indent}{context} {message}")
    elif level == logging.INFO:
        logger.info(f"{indent}{context} {message}")
    elif level == logging.WARNING:
        logger.warning(f"{indent}{context} {message}")
    elif level == logging.ERROR:
        logger.error(f"{indent}{context} {message}")
    elif level == logging.CRITICAL:
        logger.critical(f"{indent}{context} {message}")

if __name__ == "__main__":
    xlsx_file_path = "OS Info -  KIOS PUBLICATIONS FOR 2017-2018-2019-2020-2021_2022_2023_2024.xlsx"
    # xlsx_file_path = "test.xlsx"
    sheet_names = ["YEAR 2024", "YEAR 2023", "YEAR 2022"]  # Example: Add multiple sheets to process
    if not os.path.exists(xlsx_file_path):
        logger.error(f"XLSX file not found: {xlsx_file_path}")
    else:
        update_xlsx_with_zenodo_links(xlsx_file_path, sheet_names)